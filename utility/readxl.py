import openpyxl


class Readxl:
    @staticmethod
    def get_row_count(filename, sheet_name):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        rows = sheet.max_row
        return rows

    @staticmethod
    def get_column_count(filename, sheet_name):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        columns = sheet.max_column
        return columns

    @staticmethod
    def read_data(filename, sheet_name, row, column):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        data = sheet.cell(row=row, column=column).value
        return data

    @staticmethod
    def write_data(filename, sheet_name, row, column, value):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=column).value = value
        workbook.save(filename)
        workbook.close()





